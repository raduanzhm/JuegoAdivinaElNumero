import openpyxl
from openpyxl import Workbook
import matplotlib.pyplot as plt

def mostrar_resumen_y_grafico():
    try:
        wb = openpyxl.load_workbook('EstadisticaJuego.xlsx')
        hoja = wb['Partidas']
    except:
        print('No se encontró el archivo EstadisticaJuego.xlsx')
        return

    print('\n---ESTADISTICAS DE PARTIDAS---')
    print('Nombre          Dificultad     Intentos   Resultado')
    print('---------------------------------------------------')

    victorias = {}

    for fila in hoja.iter_rows(min_row=2, values_only=True):
        if not fila:
            continue
        nombre = fila[0] or 'Anónimo'
        dificultad = fila[1] or ''
        intentos = fila[2] or 0
        resultado = fila[3] or ''

        print(f'{nombre:<15}{dificultad:<15}{intentos:<10}{resultado:<10}')

        if str(resultado).lower().startswith('gan'):
            victorias[nombre] = victorias.get(nombre, 0) + 1

    print('---------------------------------------------------')

    if not victorias:
        print('No hay victorias registradas.')
        wb.close()
        return

    # Ranking simple
    ranking = sorted(victorias.items(), key=lambda x: x[1], reverse=True)

    print('\n--- ESTADISTICAS DE JUGADORES ---')
    print('Jugador         Victorias')
    print('----------------------------')
    total_victorias = 0
    for nombre, wins in ranking:
        total_victorias += wins
        print(f'{nombre:<15}{wins:>10}')
    print('----------------------------')
    print('Total jugadores:', len(ranking), ' Total victorias:', total_victorias)

    nombres = [p[0] for p in ranking]
    valores = [p[1] for p in ranking]
    plt.bar(nombres, valores)
    plt.title('Victorias por Jugador')
    plt.xlabel('Jugador')
    plt.ylabel('Victorias')
    plt.xticks(rotation=45, ha='right')
    plt.tight_layout()
    plt.show()

    wb.close()


def guardar_partida(nombre, modo, dificultad, intentos_permitidos, intentos_usados, gano, numero_secreto):
    ruta = 'EstadisticaJuego.xlsx'
    nombre_hoja = 'Partidas'

  
    try:
        wb = openpyxl.load_workbook(ruta)
        if nombre_hoja in wb.sheetnames:
            hoja = wb[nombre_hoja]
        else:
            hoja = wb.create_sheet(nombre_hoja)
            hoja.append(['Nombre', 'Dificultad', 'Intentos', 'Resultado'])
    except:
        wb = Workbook()
        hoja = wb.active
        hoja.title = nombre_hoja
        hoja.append(['Nombre', 'Dificultad', 'Intentos', 'Resultado'])

    nombre = (nombre or '').strip() or 'Anónimo'
    resultado = 'Ganó' if gano else 'Perdió'
    hoja.append([nombre, dificultad, intentos_usados, resultado])

    wb.save(ruta)
    wb.close()









